"""
Aplana el Excel de analisis fundamental de ETFs (Dashboard + una hoja "Ficha"
por ETF) a JSON que la app Streamlit lee en runtime, sin depender de openpyxl
ni del .xlsx en produccion. Espejo de export_excel.py pero para ETFs, que no
tienen DCF/WACC (valoracion via Costo Total Real y percentil historico en vez
de Margen de Seguridad).

Uso:
    python scripts/export_excel_etfs.py "C:\\ruta\\a\\Analisis_Fundamental_ETFs.xlsx"

Genera:
    data/etfs_ranking.json     - resumen Dashboard, 1 registro por ETF
    data/etfs/<TICKER>.json    - Ficha completa por ETF

No commitear nunca el .xlsx fuente, solo estos JSON derivados.
"""
import datetime
import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = REPO_ROOT / "data"
ETFS_DIR = DATA_DIR / "etfs"

NON_ETF_SHEETS = {"Dashboard", "Ficha Detallada (plantilla)", "Metodologia", "Glosario"}

DASHBOARD_COLUMNS = {
    "nombre": 1, "ticker": 2, "categoria": 3, "emisor": 4, "fecha_analisis": 5,
    "indice_benchmark": 6, "aum_b": 7, "ter_pct": 8, "antiguedad_anios": 9,
    "estructura_replica": 10, "domicilio": 11, "vol_diario_m": 12,
    "bid_ask_spread_pct": 13, "n_holdings": 14, "top10_holdings_pct": 15,
    "tracking_diff_pct": 16, "tracking_error_pct": 17, "rent_1y_pct": 18,
    "rent_3y_pct": 19, "rent_5y_pct": 20, "rent_10y_pct": 21,
    "volatilidad_pct": 22, "max_drawdown_pct": 23, "sharpe_ratio": 24,
    "beta_sp500": 25, "dividend_yield_pct": 26, "frecuencia_distribucion": 27,
    "pe_ponderado": 28, "pb_ponderado": 29, "duracion_efectiva_anios": 30,
    "ytm_pct": 31, "calidad_crediticia": 32, "costo_total_real_pct": 33,
    "premium_discount_pct": 34, "rol_portafolio": 35, "riesgo_principal": 36,
    "catalizador_principal": 37, "score_total": 38, "decision_full": 39,
    "notas_valoracion": 40,
}

# Campos que en el Excel vienen como fraccion (0.0006 = 0.06%) y se
# normalizan a puntos porcentuales para mostrar directo en la UI.
PCT_FIELDS = {
    "ter_pct", "top10_holdings_pct", "tracking_diff_pct", "tracking_error_pct",
    "rent_1y_pct", "rent_3y_pct", "rent_5y_pct", "rent_10y_pct",
    "volatilidad_pct", "max_drawdown_pct", "dividend_yield_pct",
    "ytm_pct", "costo_total_real_pct", "premium_discount_pct",
    "bid_ask_spread_pct",
}

# Decisiones observadas en el Excel de ETFs (distintas a las de acciones:
# no hay DCF, asi que no existen badges tipo "COMPRAR" por margen de
# seguridad sino por postura de portafolio).
DECISION_PATTERNS = [
    ("NO APTO", "NO APTO"),
    ("REDUCIR", "REDUCIR/VENDER"),
    ("ESPERAR MEJOR ENTRADA", "ESPERAR ENTRADA"),
    ("SATELITE TACTICO", "SATÉLITE TÁCTICO"),
    ("MEJOR PERFIL DE LA CATEGORIA", "SATÉLITE TÁCTICO"),
    ("MAYOR SCORE DE LA CATEGORIA", "SATÉLITE TÁCTICO"),
    ("MANTENER", "MANTENER"),
    ("COMPRAR", "COMPRAR"),
    ("EVITAR", "EVITAR"),
]


def _strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


def decision_badge(decision_full):
    if not decision_full:
        return "SIN CLASIFICAR"
    lead = re.split(r"\s[-–]\s|[,(]", decision_full.strip(), maxsplit=1)[0]
    lead = _strip_accents(lead).upper()
    for prefix, badge in DECISION_PATTERNS:
        if prefix in lead:
            return badge
    return "SIN CLASIFICAR"


_NA_VALUES = {"N/A", "N/D", "NA", "-", ""}


def clean(v):
    if isinstance(v, str):
        v = v.strip()
        if v.upper() in _NA_VALUES:
            return None
        return v if v else None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()[:10]
    return v


def parse_ficha(ws):
    header = {}
    for row in range(4, 9):
        label = ws.cell(row=row, column=2).value
        value = ws.cell(row=row, column=3).value
        if label:
            header[str(label).strip()] = clean(value)

    secciones = []
    current = None

    max_row = ws.max_row
    for r in range(9, max_row + 1):
        b_val = clean(ws.cell(row=r, column=2).value)
        c_val = clean(ws.cell(row=r, column=3).value)
        d_val = clean(ws.cell(row=r, column=4).value)

        if b_val and c_val is None and d_val is None:
            current = {"titulo": b_val, "aspectos": []}
            secciones.append(current)
            continue

        if current is not None and b_val and (c_val is not None or d_val is not None):
            rating = d_val if isinstance(d_val, (int, float)) else None
            current["aspectos"].append({
                "aspecto": b_val,
                "notas": c_val,
                "calificacion": rating,
            })

    # La ultima "seccion" del template ("Score promedio (1-5)") en realidad
    # solo contiene el aspecto "Decision / Tesis final": es la conclusion,
    # no una seccion evaluable mas. Se separa igual que en export_excel.py.
    conclusion = {}
    if secciones and secciones[-1]["titulo"].strip().upper().startswith("SCORE PROMEDIO"):
        ultima = secciones.pop()
        tesis_asp = next(
            (a for a in ultima["aspectos"] if "TESIS" in _strip_accents(a["aspecto"]).upper()
             or "DECISION" in _strip_accents(a["aspecto"]).upper()),
            None,
        )
        if tesis_asp:
            conclusion["tesis"] = tesis_asp["notas"]

    for sec in secciones:
        ratings = [a["calificacion"] for a in sec["aspectos"] if a["calificacion"] is not None]
        sec["rating_promedio"] = round(sum(ratings) / len(ratings), 2) if ratings else None

    return {
        "nombre": header.get("Nombre ETF"),
        "ticker": header.get("Ticker"),
        "categoria": header.get("Categoria/Clase de Activo"),
        "emisor": header.get("Emisor"),
        "fecha_analisis": str(header.get("Fecha Analisis")) if header.get("Fecha Analisis") else None,
        "secciones": secciones,
        "conclusion": conclusion,
    }


def main():
    if len(sys.argv) != 2:
        print("Uso: python scripts/export_excel_etfs.py <ruta al xlsx>")
        sys.exit(1)

    xlsx_path = Path(sys.argv[1])
    if not xlsx_path.exists():
        print(f"No se encontro el archivo: {xlsx_path}")
        sys.exit(1)

    print(f"Leyendo {xlsx_path.name} ...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    dash = wb["Dashboard"]
    sheet_names = set(wb.sheetnames)

    ranking = []
    warnings = []
    ETFS_DIR.mkdir(parents=True, exist_ok=True)

    row = 5
    while row <= dash.max_row:
        nombre = dash.cell(row=row, column=DASHBOARD_COLUMNS["nombre"]).value
        if not nombre:
            row += 1
            continue

        rec = {}
        for key, col in DASHBOARD_COLUMNS.items():
            rec[key] = clean(dash.cell(row=row, column=col).value)

        ticker = rec["ticker"]
        if not ticker:
            row += 1
            continue

        for pct_field in PCT_FIELDS:
            val = rec.get(pct_field)
            if isinstance(val, (int, float)):
                rec[pct_field] = round(val * 100, 3)

        decision_full = rec.get("decision_full")
        rec["decision_badge"] = decision_badge(decision_full)
        rec["empresa"] = rec["nombre"]  # alias para reusar helpers de UI compartidos

        has_ficha = ticker in sheet_names and ticker not in NON_ETF_SHEETS
        rec["has_ficha"] = has_ficha
        if not has_ficha:
            warnings.append(f"Sin hoja de Ficha para ticker de Dashboard: {ticker} (fila {row})")

        ranking.append(rec)

        if has_ficha:
            ficha = parse_ficha(wb[ticker])
            with open(ETFS_DIR / f"{ticker}.json", "w", encoding="utf-8") as f:
                json.dump(ficha, f, ensure_ascii=False, indent=2)

        row += 1

    extra_sheets = [
        s for s in wb.sheetnames
        if s not in NON_ETF_SHEETS and s not in {r["ticker"] for r in ranking}
    ]
    for s in extra_sheets:
        warnings.append(f"Hoja de Ficha sin fila correspondiente en Dashboard: {s}")

    ranking.sort(key=lambda r: r["ticker"])
    ETFS_DIR.mkdir(parents=True, exist_ok=True)
    with open(DATA_DIR / "etfs_ranking.json", "w", encoding="utf-8") as f:
        json.dump(ranking, f, ensure_ascii=False, indent=2)

    print(f"\nOK: {len(ranking)} ETFs en data/etfs_ranking.json")
    print(f"OK: {sum(1 for r in ranking if r['has_ficha'])} fichas en data/etfs/")

    if warnings:
        print(f"\n⚠ {len(warnings)} advertencias:")
        for w in warnings:
            print(f"  - {w}")


if __name__ == "__main__":
    main()
