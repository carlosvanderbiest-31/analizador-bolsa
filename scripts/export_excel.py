"""
Aplana el Excel de analisis fundamental (Dashboard + una hoja "Ficha" por
empresa) a JSON que la app Streamlit lee en runtime, sin depender de openpyxl
ni del .xlsx en produccion.

Uso:
    python scripts/export_excel.py "C:\\ruta\\a\\Analisis_Fundamental_Empresas_ nuevo (version 1).xlsx"

Genera:
    data/ranking.json              - resumen Dashboard, 1 registro por empresa
    data/companies/<TICKER>.json   - Ficha completa (10 secciones) por empresa

No commitear nunca el .xlsx fuente, solo estos JSON derivados.
"""
import datetime
import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = REPO_ROOT / "data"
COMPANIES_DIR = DATA_DIR / "companies"

# Alias conocidos: ticker del Dashboard -> nombre real de la hoja de Ficha
# (Excel no permite puntos en nombres de hoja, ej. BRK.B se guarda como BRK)
SHEET_ALIASES = {
    "BRK.B": "BRK",
}

NON_COMPANY_SHEETS = {"Dashboard", "Ficha Detallada (plantilla)", "Glosario"}

DASHBOARD_COLUMNS = {
    "empresa": 1, "ticker": 2, "sector": 3, "fecha_analisis": 4,
    "modelo_negocio_resumen": 5, "clientes": 6, "competidores": 7, "proveedores": 8,
    "moat": 9, "riesgos": 10, "calif_management": 11,
    "deuda_neta_ebitda": 12, "razon_corriente": 13,
    "margen_bruto": 14, "margen_operativo": 15, "margen_neto": 16,
    "roe": 17, "roic": 18, "cagr_ingresos_5y": 19, "fcf_yield": 20, "capex_ventas": 21,
    "pe": 22, "ev_ebitda": 23, "p_fcf": 24,
    "precio_actual": 25, "valor_intrinseco": 26, "margen_seguridad_raw": 27,
    "riesgo_principal": 28, "catalizador_principal": 29,
    "score_total": 30, "decision_full": 31, "notas": 32,
    "rotacion_activos": 33, "apalancamiento_financiero": 34, "roe_dupont": 35,
}

DECISION_PATTERNS = [
    ("NO APTO", "ESPECULATIVO"),
    ("ESPECULATIVO", "ESPECULATIVO"),
    ("SITUACION ESPECIAL", "SITUACION ESPECIAL"),
    ("NO COMPRAR", "NO COMPRAR"),
    ("NO ABRIR", "MANTENER"),
    ("EVITAR", "EVITAR"),
    ("VENDER", "VENDER"),
    ("REDUCIR", "REDUCIR"),
    ("COMPRAR", "COMPRAR"),
    ("MANTENER", "MANTENER"),
]


def _strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


def decision_badge(decision_full):
    if not decision_full:
        return "SIN CLASIFICAR"
    lead = re.split(r"\s[-–]\s|[,(]", decision_full.strip(), maxsplit=1)[0]
    lead = _strip_accents(lead).upper()
    cautela = "CAUTELA" in _strip_accents(decision_full).upper()[:80]
    for prefix, badge in DECISION_PATTERNS:
        if prefix in lead:
            if badge in ("COMPRAR", "MANTENER") and cautela:
                return f"{badge} (CAUTELA)"
            return badge
    return "SIN CLASIFICAR"


_NA_VALUES = {"N/A", "N/D", "NA", "-", ""}

# Casos verificados a mano (leyendo la nota de la Ficha) donde la heuristica
# numerica de abajo da un resultado incorrecto para Margen de Seguridad,
# porque el valor real ya viene en puntos porcentuales pero es < 3 en
# magnitud absoluta (indistinguible de una fraccion por el numero solo).
MOS_OVERRIDES = {
    "BRK.B": -0.5,  # nota de la Ficha dice explicitamente "-0.5%"
}


def clean(v):
    if isinstance(v, str):
        v = v.strip()
        if v.upper() in _NA_VALUES:
            return None
        return v if v else None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()[:10]
    return v


def normalize_pct(raw):
    """Los campos porcentuales del Dashboard vienen en formato mixto fila a
    fila: a veces fraccion (0.0088734 = 0.89%, tipico de una formula), a
    veces ya en puntos porcentuales (78.47 = +78.47%, o 2.84 = +2.84%,
    tipico de un valor tipeado a mano), a veces string con '%' pegado
    ("114.29%").

    Heuristica: la MAGNITUD sola no alcanza (valores tipeados a mano
    como 1.24 o 2.84 son ambiguos con una fraccion). La senial confiable
    es la PRECISION DECIMAL: los valores de formula tienen cola decimal
    larga/irregular (0.0088734, -0.6399), mientras que los tipeados a
    mano quedan "limpios" a <=2 decimales (78.47, 2.84, -0.5) incluso
    cuando son chicos en magnitud. Verificado contra el texto de la
    Decision de varias filas (ej. KMI raw=1.24 -> Decision dice
    explicitamente "practicamente neutro (+1.2%)", NO +124%).
    Solo se trata como fraccion cuando tiene mas de 2 decimales Y su
    magnitud absoluta es chica. Devuelve (valor, es_heuristico)."""
    v = clean(raw)
    if v is None:
        return None, False
    if isinstance(v, str):
        try:
            return round(float(v.replace("%", "").replace(",", ".").strip()), 2), False
        except ValueError:
            return None, False
    try:
        v = float(v)
    except (TypeError, ValueError):
        return None, False
    tecleado_a_mano = abs(v - round(v, 2)) < 1e-6
    if not tecleado_a_mano and abs(v) < 3:
        return round(v * 100, 2), True
    return round(v, 2), False


def parse_ficha(ws):
    header = {}
    for row in range(4, 9):
        label = ws.cell(row=row, column=2).value
        value = ws.cell(row=row, column=3).value
        if label:
            header[str(label).strip()] = clean(value)

    secciones = []
    conclusion = {}
    current = None
    in_conclusion = False

    max_row = ws.max_row
    for r in range(11, max_row + 1):
        a_val = clean(ws.cell(row=r, column=1).value)
        b_val = clean(ws.cell(row=r, column=2).value)
        c_val = clean(ws.cell(row=r, column=3).value)
        d_val = clean(ws.cell(row=r, column=4).value)

        if a_val:
            if "CONCLUSION" in _strip_accents(a_val).upper():
                in_conclusion = True
                current = None
                continue
            if in_conclusion:
                continue
            current = {"titulo": a_val, "aspectos": []}
            secciones.append(current)
            continue

        if in_conclusion and b_val:
            label = _strip_accents(b_val).upper()
            if "SCORE" in label:
                conclusion["score_promedio"] = c_val if isinstance(c_val, (int, float)) else None
            elif "RECOMENDACION" in label:
                conclusion["recomendacion"] = c_val
            elif "TESIS" in label:
                conclusion["tesis"] = c_val
            continue

        if current is not None and b_val:
            rating = d_val if isinstance(d_val, (int, float)) else None
            current["aspectos"].append({
                "aspecto": b_val,
                "notas": c_val,
                "calificacion": rating,
            })

    for sec in secciones:
        ratings = [a["calificacion"] for a in sec["aspectos"] if a["calificacion"] is not None]
        sec["rating_promedio"] = round(sum(ratings) / len(ratings), 2) if ratings else None

    return {
        "empresa": header.get("Empresa"),
        "ticker": header.get("Ticker"),
        "sector": header.get("Sector / Industria"),
        "fecha_analisis": str(header.get("Fecha de análisis")) if header.get("Fecha de análisis") else None,
        "analista": header.get("Analista"),
        "secciones": secciones,
        "conclusion": conclusion,
    }


def main():
    if len(sys.argv) != 2:
        print("Uso: python scripts/export_excel.py <ruta al xlsx>")
        sys.exit(1)

    xlsx_path = Path(sys.argv[1])
    if not xlsx_path.exists():
        print(f"No se encontro el archivo: {xlsx_path}")
        sys.exit(1)

    print(f"Leyendo {xlsx_path.name} ...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    dash = wb["Dashboard"]
    sheet_names = set(wb.sheetnames)

    ranking = []
    warnings = []
    unmatched_decisions = set()

    row = 5
    while row <= dash.max_row:
        empresa = dash.cell(row=row, column=DASHBOARD_COLUMNS["empresa"]).value
        if not empresa:
            row += 1
            continue

        rec = {}
        for key, col in DASHBOARD_COLUMNS.items():
            rec[key] = clean(dash.cell(row=row, column=col).value)

        ticker = rec["ticker"]
        if not ticker:
            row += 1
            continue

        raw_mos = rec.pop("margen_seguridad_raw")
        if ticker in MOS_OVERRIDES:
            rec["margen_seguridad_pct"] = MOS_OVERRIDES[ticker]
            rec["mos_approx"] = False
        else:
            mos_pct, mos_approx = normalize_pct(raw_mos)
            rec["margen_seguridad_pct"] = mos_pct
            rec["mos_approx"] = mos_approx

        for pct_field in (
            "margen_bruto", "margen_operativo", "margen_neto", "roe", "roic",
            "cagr_ingresos_5y", "fcf_yield", "capex_ventas", "roe_dupont",
        ):
            val, _ = normalize_pct(rec.get(pct_field))
            rec[pct_field] = val

        decision_full = rec.get("decision_full")
        badge = decision_badge(decision_full)
        rec["decision_badge"] = badge
        if badge == "SIN CLASIFICAR" and decision_full:
            unmatched_decisions.add(f"{ticker}: {decision_full[:60]}")

        sheet_name = SHEET_ALIASES.get(ticker, ticker)
        has_ficha = sheet_name in sheet_names and sheet_name not in NON_COMPANY_SHEETS
        rec["has_ficha"] = has_ficha
        if not has_ficha:
            warnings.append(f"Sin hoja de Ficha para ticker de Dashboard: {ticker} (fila {row})")

        ranking.append(rec)

        if has_ficha:
            ficha = parse_ficha(wb[sheet_name])
            with open(COMPANIES_DIR / f"{ticker}.json", "w", encoding="utf-8") as f:
                json.dump(ficha, f, ensure_ascii=False, indent=2)

        row += 1

    extra_sheets = [
        s for s in wb.sheetnames
        if s not in NON_COMPANY_SHEETS
        and s not in {SHEET_ALIASES.get(r["ticker"], r["ticker"]) for r in ranking}
    ]
    for s in extra_sheets:
        warnings.append(f"Hoja de Ficha sin fila correspondiente en Dashboard: {s}")

    ranking.sort(key=lambda r: r["ticker"])
    DATA_DIR.mkdir(exist_ok=True)
    with open(DATA_DIR / "ranking.json", "w", encoding="utf-8") as f:
        json.dump(ranking, f, ensure_ascii=False, indent=2)

    print(f"\nOK: {len(ranking)} empresas en data/ranking.json")
    print(f"OK: {sum(1 for r in ranking if r['has_ficha'])} fichas en data/companies/")

    if warnings:
        print(f"\n⚠ {len(warnings)} advertencias:")
        for w in warnings:
            print(f"  - {w}")
    if unmatched_decisions:
        print(f"\n⚠ {len(unmatched_decisions)} decisiones sin badge reconocido (quedan como 'SIN CLASIFICAR'):")
        for u in sorted(unmatched_decisions):
            print(f"  - {u}")


if __name__ == "__main__":
    main()
